from django.core.management.base import BaseCommand
from faker import Faker
from datetime import datetime
from django.contrib.auth.models import User
from social.models import UserProfile, Interest, Service, Event, Log, Tag
import random
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Command Information"

    def add_arguments(self, parser):
        parser.add_argument('fakedata_count', type=int)

    def handle(self, *args, **kwargs):
        fake = Faker(["tr_TR"])
        fake_en = Faker(["en_US"])
        fakedata_count = kwargs['fakedata_count']

        profile_picture_dict = {1: 'stock-female-profile-1.jpg',
                                2: 'stock-female-profile-2.jpg',
                                3: 'stock-female-profile-3.jpg',
                                4: 'stock-female-profile-4.jpg',
                                5: 'stock-female-profile-5.jpg',
                                6: 'stock-female-profile-6.jpg',
                                7: 'stock-male-profile-1.jpg',
                                8: 'stock-male-profile-2.jpg',
                                9: 'stock-male-profile-1.jpg',
                                10: 'stock-male-profile-4.jpg',
                                11: 'stock-male-profile-5.jpg',
                                12: 'stock-male-profile-6.jpg',
                                13: 'stock-male-profile-7.jpg',
                                14: 'antonettejford.jpeg',
                                15: 'jillbass.jpeg',
                                16: 'melektaner.jpeg'}

        event_picture_dict = {1: 'accounting.jpeg',
                              2: 'beancooking.jpeg',
                              3: 'eglence.png',
                              4: 'english.jpeg',
                              5: 'finance.jpeg',
                              6: 'ukulele.jpeg',
                              7: 'icli-kofte.jpg',
                              8: 'istock-free-dogwalking.jpg',
                              9: 'istock-free-horon.jpg',
                              10: 'istock-free-ney.jpg',
                              11: 'istock-free-origami.jpg',
                              12: 'istock-free-poetry.jpg',
                              13: 'istock-free-shogi.jpg',
                              14: 'istock-free-strawhat.jpg',
                              15: 'istock-free-vegan.jpg',
                              16: 'pexels-free-juggling.jpg'}

        service_picture_dict = {1: 'accounting.jpeg',
                                2: 'beancooking.jpeg',
                                3: 'eglence.png',
                                4: 'english.jpeg',
                                5: 'finance.jpeg',
                                6: 'ukulele.jpeg',
                                7: 'icli-kofte.jpg',
                                8: 'istock-free-dogwalking.jpg',
                                9: 'istock-free-horon.jpg',
                                10: 'istock-free-ney.jpg',
                                11: 'istock-free-origami.jpg',
                                12: 'istock-free-poetry.jpg',
                                13: 'istock-free-shogi.jpg',
                                14: 'istock-free-strawhat.jpg',
                                15: 'istock-free-vegan.jpg',
                                16: 'pexels-free-juggling.jpg'}

        category_dict = {1: 'Sports',
                         2: 'Board Games',
                         3: 'Literature',
                         4: 'Cooking',
                         5: 'Entertainment',
                         6: 'Music',
                         7: 'Science',
                         8: 'Technology',
                         9: 'Art',
                         10: 'Dance'}

        def load_interest_example():
            wb = load_workbook(filename='social/management/commands/interest_example.xlsx')
            ws = wb.active
            interests_dict = {}
            for row in list(ws.rows)[1:]:
                interests_dict[row[0].value] = [str(c.value) for c in row[1:]]
            return interests_dict

        def load_locations_tr():
            wb = load_workbook(filename='social/management/commands/ilceler_tr.xlsx')
            ws = wb.active
            loc_dict = {}
            for row in list(ws.rows)[1:]:
                loc_dict[row[0].value] = [str(c.value) for c in row[1:]]
            return loc_dict

        # This function returns the list of the districts of a given city
        def get_district_of_city(city_no):
            district_key_list = []
            districts = locations_dict.items()
            for district in districts:
                if district[1][0] == str(city_no):
                    district_key_list.append(district[0])
            return district_key_list

        # This function creates random users and updates the user_profiles randomly and adds an interest to that user
        def create_fake_user_data():
            name = fake.name()
            if len(name) > 30:
                name = name[0:30]
            bio = fake_en.paragraph(nb_sentences=5, variable_nb_sentences=False)
            birth_date = datetime.combine(fake.date_of_birth(minimum_age=18, maximum_age=75), datetime.min.time())
            city_no = random.randint(1, 81)
            min_district_key = min(get_district_of_city(city_no))
            max_district_key = max(get_district_of_city(city_no))
            district = random.randint(min_district_key, max_district_key)
            location = locations_dict[district][2] + ',' + locations_dict[district][3]
            password = 'admin123+'
            last_login = fake.date_time_between_dates(datetime_start=datetime(2022, 5, 20),
                                                      datetime_end=datetime(2022, 5, 22))
            username = fake.user_name()
            while len(User.objects.filter(username=username)) > 0:
                username = fake.user_name()
            email = fake.email()
            date_joined = fake.date_time_between_dates(datetime_start=datetime(2022, 1, 1),
                                                       datetime_end=datetime(2022, 5, 20))

            user = User.objects.create_user(username=username, email=email, password=password, is_superuser=False,
                                            is_staff=False, is_active=True, date_joined=date_joined,
                                            last_login=last_login)
            profile = UserProfile(user.id)
            profile.name = name
            profile.location = location
            profile.bio = bio
            picture_no = random.randint(1, 16)
            profile.picture = 'uploads/profile_pictures/' + profile_picture_dict[picture_no]
            profile.birth_date = birth_date
            profile.save()

            interest_count = random.randint(1, 5)
            arr = random.sample(range(1, len(interest_dict)), interest_count)
            for i in range(len(arr)):
                interest_no = arr[i]
                new_interest = Interest.objects.create(user=user, name=interest_dict[interest_no][0],
                                                       wiki_description=interest_dict[interest_no][1], implicit=False,
                                                       origin='profile')
                new_interest.save()

        # This function makes random users follow other users
        def create_fake_following_data():
            last_user = User.objects.latest('id')
            arr = random.sample(range(1, last_user.id-1), 2)
            user1_id = arr[0]
            user2_id = arr[1]

            user1 = UserProfile.objects.get(pk=user1_id)
            user1.followers.add(user2_id)

        # This function creates the categories to be used in the service creation
        def create_categories():
            for i in range(len(category_dict)):
                category = category_dict[i+1]
                if len(Tag.objects.filter(tag=category)) == 0:
                    last_user = User.objects.latest('id')
                    tag = Tag.objects.create(tag=category, requester_id=last_user.id)
                    tag.save()

        # This function creates random services and their service creation logs
        def create_fake_service_data():
            name = fake_en.sentence()
            sentence_count = random.randint(1, 7)
            desc = fake_en.paragraph(nb_sentences=sentence_count, variable_nb_sentences=False)
            city_no = random.randint(1, 81)
            min_district_key = min(get_district_of_city(city_no))
            max_district_key = max(get_district_of_city(city_no))
            district = random.randint(min_district_key, max_district_key)
            location = locations_dict[district][2] + ',' + locations_dict[district][3]
            city = locations_dict[district][1]
            create_date = fake.date_time_between_dates(datetime_start=datetime(2022, 1, 1),
                                                       datetime_end=datetime(2022, 5, 20))
            last_user = User.objects.latest('id')
            user_id = random.randint(1, last_user.id)
            capacity = random.randint(1, 5)
            duration = random.randint(1, 3)
            service_date = fake.future_datetime(end_date="+60d")
            wiki_no = random.randint(1, 50)
            wiki_desc = interest_dict[wiki_no][0] + ' as a(n) ' + interest_dict[wiki_no][1]
            picture_no = random.randint(1, 16)
            picture = 'uploads/service_pictures/' + service_picture_dict[picture_no]
            category_id = random.randint(1, Tag.objects.latest('id').id)

            new_service = Service.objects.create(createddate=create_date, description=desc, creater_id=user_id,
                                                 capacity=capacity, duration=duration, location=location, name=name,
                                                 servicedate=service_date, wiki_description=wiki_desc, city=city,
                                                 picture=picture, category_id=category_id)
            new_service.save()

            user = User(user_id)
            log = Log.objects.create(operation="createservice", itemType="service", itemId=new_service.pk,
                                     userId=user, date=create_date)
            log.save()

        # This function creates random events and their event creation logs
        def create_fake_event_data():
            name = fake_en.sentence()
            sentence_count = random.randint(1, 7)
            desc = fake_en.paragraph(nb_sentences=sentence_count, variable_nb_sentences=False)
            city_no = random.randint(1, 81)
            min_district_key = min(get_district_of_city(city_no))
            max_district_key = max(get_district_of_city(city_no))
            district = random.randint(min_district_key, max_district_key)
            location = locations_dict[district][2] + ',' + locations_dict[district][3]
            city = locations_dict[district][1]
            create_date = fake.date_time_between_dates(datetime_start=datetime(2022, 1, 1),
                                                       datetime_end=datetime(2022, 5, 20))
            last_user = User.objects.latest('id')
            user_id = random.randint(1, last_user.id)
            capacity = random.randint(5, 20)
            duration = random.randint(2, 6)
            event_date = fake.future_datetime(end_date="+60d")
            wiki_no = random.randint(1, 50)
            wiki_desc = interest_dict[wiki_no][0] + ' as a(n) ' + interest_dict[wiki_no][1]
            picture_no = random.randint(1, 16)
            picture = 'uploads/event_pictures/' + event_picture_dict[picture_no]

            new_event = Event.objects.create(eventcreateddate=create_date, eventname=name, eventdescription=desc,
                                             event_wiki_description=wiki_desc, eventlocation=location,
                                             eventdate=event_date, eventcapacity=capacity, eventduration=duration,
                                             eventcreater_id=user_id, city=city, eventpicture=picture)
            new_event.save()

            user = User(user_id)
            log = Log.objects.create(operation="createevent", itemType="event", itemId=new_event.pk,
                                     userId=user, date=create_date)
            log.save()

        if fakedata_count > 0:

            locations_dict = load_locations_tr()
            interest_dict = load_interest_example()

            for _ in range(fakedata_count):
                create_fake_user_data()

            for _ in range(fakedata_count):
                create_fake_following_data()
            
            create_categories()

            for _ in range(fakedata_count):
                create_fake_service_data()

            for _ in range(fakedata_count):
                create_fake_event_data()

            print(str(fakedata_count) + ' items are created successfully.')
        else:
            print('You need to give a positive number as an argument')
